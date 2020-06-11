#https://qiita.com/mimitaro/items/2b48b377b14018bdf6cb

from openpyxl import load_workbook

# エクセルファイルのロード
excel_path='sampleNaka1.xlsx'
wb = load_workbook(filename=excel_path, read_only=True)

# シートのロード
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']

# セルの値取得
cell_Sheet1_a1_value = sheet1['A1'].value
cell_Sheet2_a1_value = sheet2['A1'].value

# 取得した値の表示
print('A1', cell_Sheet1_a1_value)
print('B1', cell_Sheet2_a1_value)

#Sheet1のA列のサイズと、Sheet2のA列のサイズを調べる
fmt = "{:7d} {:7d} {:7d} {:7d}"
print("min_col min_row max_col max_row")
print(fmt.format(sheet1.min_column,sheet1.min_row, sheet1.max_column, sheet1.max_row))
print(fmt.format(sheet2.min_column,sheet2.min_row, sheet2.max_column, sheet2.max_row))
#

#Sheet2のA列に名前がある人のなかで、Sheet1のA列に名前がない人がいたら、その人の名前を表示する




# ロードしたExcelファイルを閉じる
wb.close()
